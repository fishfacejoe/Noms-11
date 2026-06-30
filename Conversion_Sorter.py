import pandas as pd
import json
from openpyxl import load_workbook
# python -m http.server
# localhost:8000
excel_file_name = "Noms11 (eggsdee).xlsx"

def excel_to_json(file_path, output_path):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    # Charger le fichier Excel avec pandas
    df = pd.read_excel(file_path, engine='openpyxl')

    # Vérifier que les colonnes nécessaires existent
    required_columns = {"ID", "Anime Name", "Song Type", "Song Info", "mp3 Links"}
    if not required_columns.issubset(df.columns):
        raise ValueError(f"Column name don't match these: {required_columns}")

    result = []

    for idx, row in df.iterrows():
        if pd.isna(row["ID"]):  # Ignorer les lignes où ID est vide
            continue

        # Extraire le lien hypertexte de la colonne "Song Info" (vidéo)
        song_info = row["Song Info"]
        video_link = None
        text_value = song_info  # Suppose que song_info est une chaîne de texte

        # Récupérer le lien vidéo via openpyxl si c'est un objet hyperlien
        song_info_cell = sheet.cell(row=idx + 2, column=df.columns.get_loc("Song Info") + 1)  # +2 pour ignorer l'entête
        if song_info_cell.hyperlink:
            video_link = song_info_cell.hyperlink.target

        # Extraire le lien hypertexte de la colonne "mp3 Links EU"
        mp3_link = None
        mp3_cell = sheet.cell(row=idx + 2, column=df.columns.get_loc("mp3 Links") + 1)
        if mp3_cell.hyperlink:
            mp3_link = mp3_cell.hyperlink.target

        # Construire l'objet JSON
        entry = {
            "id": int(row["ID"]),
            "anime": f"{row['Anime Name']} - {row['Song Type']}",
            "name": text_value,
            "video": video_link,
            "mp3": mp3_link
        }

        result.append(entry)

    # Sauvegarder le JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=4)

    print(f"JSON généré avec succès: {output_path}")

# Exemple d'utilisation
excel_to_json(excel_file_name, "songList.json")